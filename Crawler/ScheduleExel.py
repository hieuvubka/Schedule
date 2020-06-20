#!/usr/local/bin/python
# coding: utf-8
import pandas as pd
from Crawler.ScheduleCrawler import ScheduleCrawler
from openpyxl import load_workbook

class a():
    def __init__(self, studentCode):
        schedule = ScheduleCrawler(studentCode)
        if schedule.classCount == 0:
            pass
        else:
            self.fileName = 'ScheduleTable.xlsx'
            self.wb = load_workbook(self.fileName)
            sheet = self.wb.active

            if(sheet['A1'] != 'Mã SV'):
                sheet['A1'] = 'Mã SV'
                sheet['B1'] = 'Thời gian'
                sheet['C1'] = 'Tuần học'
                sheet['D1'] = 'Phòng học'
                sheet['E1'] = 'Mã lớp'
                sheet['F1'] = 'Loại lớp'
                sheet['G1'] = 'Nhóm'
                sheet['H1'] = 'Mã HP'
                sheet['I1'] = 'Tên lớp'
                sheet['J1'] = 'Ghi chú'

            check = 0
            df = pd.read_excel(self.fileName)
            check = df['Mã SV'].where(df['Mã SV'] == schedule.student_code)

            if len(check.dropna()) > 0:
                pass
            else:
                row = sheet.max_row
                for j in range( schedule.classCount):
                    sheet['A%d' % (j + row)] = schedule.student_code
                    sheet['B%d' % (j + row)] = schedule.time[j]
                    sheet['C%d' % (j + row)] = schedule.week[j]
                    sheet['D%d' % (j + row)] = schedule.room[j]
                    sheet['E%d' % (j + row)] = schedule.classId[j]
                    sheet['F%d' % (j + row)] = schedule.classType[j]
                    sheet['G%d' % (j + row)] = schedule.group[j]
                    sheet['H%d' % (j + row)] = schedule.subjectId[j]
                    sheet['I%d' % (j + row)] = schedule.className[j]
                    sheet['J%d' % (j + row)] = schedule.note[j]

            self.wb.save(self.fileName)





